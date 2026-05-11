# Tämä tiedosto sisältää yhteiset apufunktiot työterveyskustannusdatan käsittelyyn.
#
# Käyttö:
# - Excel-tiedostojen sheetien nimissä pitää näkyä kuukausi ja vuosi
#   esimerkiksi: tammikuu2023.pdf, lokakuu2024 tai syyskuu2025
# - Kuukauden loppusumman rivillä pitää olla:
#   A-sarakkeessa "Yhteensä:"
#   B-sarakkeessa kuukauden summa
# - Datarivien sarakkeiden oletetaan olevan samassa rakenteessa kuin nykyisissä
#   tamperetaloYYYY.xlsx-tiedostoissa
#
# Uuden vuoden lisääminen:
# - Lisää uusi tamperetaloYYYY.xlsx-tiedosto samaan kansioon
# - Jos sheetien nimet ovat samassa muodossa kuin aiemmin, helpers.py toimii ilman muutoksia
#
# Jos palvelukategorioita tai laboratoriotunnistusta halutaan laajentaa:
# - muokkaa classify_category()-funktion avainsanoja
# - muokkaa LAB_CODE_ALLOWLIST- ja LAB_DESC_KEYWORDS-listoja


from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# Kuukausikartta ilman kovakoodattua vuotta.
# Vuosi luetaan automaattisesti sheetin nimestä.
MONTH_MAP = {
    "tammikuu": "01",
    "helmikuu": "02",
    "maaliskuu": "03",
    "huhtikuu": "04",
    "toukokuu": "05",
    "kesakuu": "06",
    "heinakuu": "07",
    "elokuu": "08",
    "syyskuu": "09",
    "lokakuu": "10",
    "marraskuu": "11",
    "joulukuu": "12",
}

# Laboratoriokoodit, jotka halutaan tunnistaa varmasti labroiksi.
LAB_CODE_ALLOWLIST = {
    "1026", "1046", "1078", "1128", "1137", "1185", "1189", "1216",
    "1270", "1395", "1468", "1471", "1489", "1784", "1881P", "2001",
    "2143", "2203", "2245", "2303", "2360", "2382", "2455", "2474",
    "2682", "2703", "2735", "2775", "2832", "2836", "3442", "3509",
    "3642", "3695", "3696", "3950II", "4043I", "4054", "4113P", "4224",
    "4511II", "4764", "4803", "4816", "5017", "6128", "6141", "6307",
    "6354", "6434", "650", "8022", "9100", "9828", "9951", "10143",
}

# Avainsanat, joilla laboratoriorivejä voidaan tunnistaa myös tekstin perusteella.
LAB_DESC_KEYWORDS = [
    "s-", "s -", "b-", "b -", "p-", "p -", "u-", "u -",
    "fp-", "fp -", "fs-", "fs -",
    "gluk", "hba1c", "ferrit", "psa", "t4", "tsh",
    "alat", "krea", "pvk", "lipid", "crp", "vit", "testo",
    "amy", "calpro", "para", "tvk", "bilir",
]


def normalize_text(value) -> str:
    """Normalisoi tekstin vertailua varten."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def to_float(value):
    """Muuttaa arvon liukuluvuksi, jos mahdollista."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None

    text = str(value).strip().replace(",", ".")
    text = text.replace("\u00a0", "").replace(" ", "").replace("..", ".")

    if text == "" or text.lower() == "nan":
        return None

    try:
        return float(text)
    except ValueError:
        return None


def parse_month_from_sheet(sheet_name: str) -> str:
    """
    Muodostaa sheetin nimestä arvon muotoon YYYY-MM.

    Esimerkkejä:
    - tammikuu2023.pdf -> 2023-01
    - lokakuu2024 -> 2024-10
    - syyskuu2025.pdf -> 2025-09
    """
    normalized = normalize_text(sheet_name)

    year_match = re.search(r"20\d{2}", normalized)
    if not year_match:
        raise ValueError(f"Vuotta ei löytynyt sheetin nimestä: {sheet_name}")
    year = year_match.group(0)

    for month_name, month_num in MONTH_MAP.items():
        if normalized.startswith(month_name):
            return f"{year}-{month_num}"

    raise ValueError(f"Tuntematon kuukausi: {sheet_name}")


def find_month_total_from_worksheet(ws, sheet_name: str) -> float:
    """
    Hakee kuukauden loppusumman riviltä, jossa A-sarakkeessa on 'Yhteensä:'
    ja B-sarakkeessa summa.
    """
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if len(values) < 2:
            continue

        first_cell = normalize_text(values[0])
        second_cell = to_float(values[1])

        if "yhteensa" in first_cell and second_cell is not None:
            return second_cell

    raise ValueError(
        f"Yhteensä-riviä ei löytynyt sheetiltä {sheet_name}. "
        "Varmista että A-sarakkeessa on 'Yhteensä:' ja B-sarakkeessa summa."
    )


def is_data_row(values) -> bool:
    """
    Tunnistaa, onko kyseessä varsinainen datarivi.
    Sulkee pois otsikot, yhteensä-rivit ja tyhjät rivit.
    """
    c0 = normalize_text(values[0] if len(values) > 0 else "")
    c1 = normalize_text(values[1] if len(values) > 1 else "")
    c2 = normalize_text(values[2] if len(values) > 2 else "")
    joined = f"{c0} {c1} {c2}"

    if not any([c0, c1, c2]):
        return False
    if c0.startswith("koodi"):
        return False
    if c0 == "0":
        return False
    if "yhteensa" in joined:
        return False
    if "seloste" in joined and "summa" in joined:
        return False
    if "yleismaksu ajalta" in joined:
        return False

    return True


def get_amount_from_values(values):
    """
    Hakee kustannussumman riviltä.
    Ensin yritetään sarake I (indeksi 8), sitten H (indeksi 7).
    """
    if len(values) > 8:
        amount = to_float(values[8])
        if amount is not None:
            return amount

    if len(values) > 7:
        amount = to_float(values[7])
        if amount is not None:
            return amount

    return None


def read_clean_rows(excel_path: str | Path) -> pd.DataFrame:
    """
    Lukee yhden tamperetaloYYYY.xlsx-tiedoston kaikki sheetit
    ja palauttaa siistit datarivit yhtenä DataFramena.
    """
    path = Path(excel_path)
    wb = load_workbook(path, data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        month = parse_month_from_sheet(sheet_name)

        for row in ws.iter_rows(values_only=True):
            values = list(row)

            if not is_data_row(values):
                continue

            amount = get_amount_from_values(values)
            if amount is None:
                continue

            rows.append(
                {
                    "Kuukausi": month,
                    "Koodi": "" if len(values) <= 0 or values[0] is None else str(values[0]).strip(),
                    "Kl": "" if len(values) <= 1 or values[1] is None else str(values[1]).strip(),
                    "Seloste": "" if len(values) <= 2 or values[2] is None else str(values[2]).strip(),
                    "a_hinta": to_float(values[3]) if len(values) > 3 else None,
                    "Alennus": to_float(values[4]) if len(values) > 4 else None,
                    "Hinta": to_float(values[5]) if len(values) > 5 else None,
                    "Tunnit": "" if len(values) <= 6 or values[6] is None else str(values[6]).strip(),
                    "Lkm": to_float(values[7]) if len(values) > 7 else None,
                    "Summa €": amount,
                }
            )

    df = pd.DataFrame(rows)

    if not df.empty:
        if "Summa €" in df.columns:
            df["Summa €"] = df["Summa €"].fillna(0.0)
        if "Lkm" in df.columns:
            df["Lkm"] = df["Lkm"].fillna(0.0)

    return df


def classify_category(desc: str, code: str) -> str:
    """
    Luokittelee palvelurivin pääkategoriaan.
    Näitä voi tarvittaessa täydentää myöhemmin uusilla avainsanoilla.
    """
    text = normalize_text(f"{desc} {code}")

    if any(x in text for x in ["tpsy", "psyk", "terapia", "mielenter"]):
        return "Mielenterveys"

    if any(x in text for x in ["puhelin", "sposti", "sahkoposti", "neuvonta", "ohjaus", "eta"]):
        return "Etäpalvelut ja ohjaus"

    if any(
        x in text
        for x in [
            "s-", "s -", "b-", "b -", "p-", "p -", "u-", "u -",
            "gluk", "hba1c", "ferrit", "psa", "t4", "tsh",
            "alat", "krea", "pvk", "lipid", "crp",
        ]
    ):
        return "Laboratorio"

    if any(x in text for x in ["ekg", "tahystys", "rontgen", "kuvaus", "ultra", "magneetti"]):
        return "Tutkimukset ja toimenpiteet"

    if any(x in text for x in ["hoitaja", "tyofysioterapeutti", "tth"]):
        return "Hoitaja ja fysioterapia"

    if any(x in text for x in ["laakari", "erikois"]):
        return "Lääkäripalvelut"

    if any(x in text for x in ["kanta", "yleismaksu", "raportointi", "toimintasuunnitelma", "lausunto", "selvitys tyopaikalle"]):
        return "Hallinto ja työpaikkayhteistyö"

    return "Muut"


def is_lab_row(desc: str, code: str) -> bool:
    """
    Tunnistaa laboratoriorivit joko koodin tai selosteen perusteella.
    """
    code_text = normalize_text(code).upper().replace(" ", "")
    desc_text = normalize_text(desc)

    if code_text in {c.upper() for c in LAB_CODE_ALLOWLIST}:
        return True

    return any(keyword in desc_text for keyword in LAB_DESC_KEYWORDS)


